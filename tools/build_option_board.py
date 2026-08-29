"""Convert the newest 옵션전광판 Excel export into a small JSON the site can load.

The exports/ folder is gitignored and lives only on the local PC, so the
deployed site can't read the .xlsx files directly. This script pulls the
latest snapshot into assets/option-board.json, which does get committed —
that file is what the Strategy Builder page fetches.

The bid/ask ladder in these exports is not available from the KRX Open API
(which publishes closing prices only), so this is the sole source for it.

Usage:  python tools/build_option_board.py
"""

import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
EXPORTS = ROOT / "exports"
OUT = ROOT / "assets" / "option-board.json"

# Columns in the export, 0-indexed. The sheet also carries market-maker
# quotes and a block of per-strike analytics that the page doesn't use.
COL_CALL_BID, COL_CALL_CUR, COL_CALL_ASK = 1, 2, 3
COL_STRIKE = 5
COL_PUT_BID, COL_PUT_CUR, COL_PUT_ASK = 7, 8, 9

# Filenames look like 옵션전광판_20260828_15h25m.xlsx, or 옵션전광판_20260825.xlsx
STAMP = re.compile(r"_(\d{8})(?:_(\d{1,2})h(?:(\d{1,2})m)?)?\.xlsx$")


def parse_stamp(path):
    """Sort key + display timestamp from the filename, falling back to mtime."""
    m = STAMP.search(path.name)
    if not m:
        return None
    day, hour, minute = m.group(1), m.group(2), m.group(3)
    return datetime(
        int(day[0:4]), int(day[4:6]), int(day[6:8]),
        int(hour) if hour else 0,
        int(minute) if minute else 0,
    )


def second_thursday(year, month):
    d = date(year, month, 1)
    thursdays = 0
    while True:
        if d.weekday() == 3:
            thursdays += 1
            if thursdays == 2:
                return d
        d += timedelta(days=1)


def front_month(snapshot_day):
    """KOSPI 200 options expire on the second Thursday; roll once that passes."""
    y, m = snapshot_day.year, snapshot_day.month
    if snapshot_day > second_thursday(y, m):
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return f"{y}{m:02d}"


def num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    return None


def main():
    if not EXPORTS.is_dir():
        sys.exit(f"No exports folder at {EXPORTS}")

    # "~$" files are Excel's lock files, not real workbooks.
    candidates = [
        p for p in EXPORTS.glob("옵션전광판_*.xlsx")
        if not p.name.startswith("~$") and parse_stamp(p)
    ]
    if not candidates:
        sys.exit("No 옵션전광판_*.xlsx snapshots found in exports/")

    latest = max(candidates, key=parse_stamp)
    stamp = parse_stamp(latest)

    wb = openpyxl.load_workbook(latest, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    # The two reference prices sit in the header row, above the strike
    # column and above the put Current P column.
    underlying = num(header[COL_STRIKE])
    index_price = num(header[COL_PUT_CUR])

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        strike = num(r[COL_STRIKE])
        if strike is None:
            continue
        rows.append({
            "k": strike,
            "cb": num(r[COL_CALL_BID]),
            "cc": num(r[COL_CALL_CUR]),
            "ca": num(r[COL_CALL_ASK]),
            "pb": num(r[COL_PUT_BID]),
            "pc": num(r[COL_PUT_CUR]),
            "pa": num(r[COL_PUT_ASK]),
        })
    wb.close()

    rows.sort(key=lambda x: x["k"], reverse=True)

    expiry = front_month(stamp.date())
    expiry_date = second_thursday(int(expiry[:4]), int(expiry[4:]))
    days_to_expiry = (expiry_date - stamp.date()).days

    payload = {
        "source": latest.name,
        "snapshot": stamp.strftime("%Y-%m-%dT%H:%M:00+09:00"),
        "expiry": expiry,
        "daysToExpiry": days_to_expiry,
        "underlying": underlying,
        "indexPrice": index_price,
        # KOSPI 200 options are 250,000 KRW per index point.
        "multiplier": 250000,
        "strikes": rows,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")

    size_kb = OUT.stat().st_size / 1024
    print(f"source     {latest.name}")
    print(f"snapshot   {payload['snapshot']}")
    print(f"expiry     {payload['expiry']}  ({days_to_expiry} days)")
    print(f"underlying {underlying}   index {index_price}")
    print(f"strikes    {len(rows)}")
    print(f"wrote      {OUT.relative_to(ROOT)}  ({size_kb:.0f} KB)")


if __name__ == "__main__":
    main()
